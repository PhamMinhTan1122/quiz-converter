import re
from openpyxl.styles import Font
import openpyxl
import os

OUTPUT_PATH = os.path.join(os.getcwd(), "output")


def CHECK_DIRECTORY():
    if not os.path.exists(OUTPUT_PATH):
        os.mkdir(OUTPUT_PATH)

def WRITE_NEW_EXCEL(_XLSX_FILENAME):
    CHECK_DIRECTORY()
    # create new file xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Create a Quiz"
    FONT = Font(name='Arial', bold=True, size=11)
    headers = {
            'A1': 'Question Text',
            'B1': 'Question Type',
            'C1': 'Option 1',
            'D1': 'Option 2',
            'E1': 'Option 3',
            'F1': 'Option 4',
            'G1': 'Correct Answer',
            'H1': 'Time in seconds',
            'A2': 'Text of the question (required)',
            'B2': 'Question Type',
            'C2': 'Text for option 1',
            'D2': 'Text for option 2',
            'E2': 'Text for option 3',
            'F2': 'Text for option 4',
            'G2': 'Correct Answer',
            'H2': 'Time in seconds'
        }
    for cell, header_text in headers.items():
        ws[cell] = header_text
        ws[cell].font = FONT
    match = re.search(r'([^\\]+)\\(?:\s*)([^\\]+)$', _XLSX_FILENAME)  # Modified regex pattern

    if match:
        folder_name = match.group(1) 
        clean_file_name = match.group(2).replace(".docx", ".xlsx")
        full_output_path = os.path.join(OUTPUT_PATH, folder_name) 
        if not os.path.exists(full_output_path):
            os.makedirs(full_output_path)
        full_file_path = os.path.join(full_output_path, clean_file_name)
        if os.path.exists(full_file_path):
            os.remove(full_file_path)
        wb.save(full_file_path)
        return full_file_path
    
    # wb.save(full_output_path)
    
    # return full_output_path

