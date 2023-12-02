# This module contains a function to write data to an excel file
import openpyxl
import os
from module.extractor import EXTRA_OPTIONS
from module.extractor import EXTRA_QUESTIONS
from module.empty_cell import CHECK_EMPTY_CELL


def WRITE_TO_EXCEL(data_list, _XLSX_FILENAME):
    try:
        # Load the existing workbook
        wb = openpyxl.load_workbook(_XLSX_FILENAME)
        ws = wb.active
        # Initialize a row counter
        ROW = 3

        # Loop over the data list with a step of 5
        for i in range(0, len(data_list), 5):
            # Check if the data_list has at least 4 elements from index i
            if len(data_list) >= i + 4:
                # Write the question and its options to the worksheet cells
                ws[f"A{ROW}"] = EXTRA_QUESTIONS(data_list, i) # The question
                ws[f"B{ROW}"] = "Multiple Choice"
                ws[f"C{ROW}"] = EXTRA_OPTIONS(data_list, i + 1)# Option A
                ws[f"D{ROW}"] = EXTRA_OPTIONS(data_list, i + 2) # Option B
                ws[f"E{ROW}"] = EXTRA_OPTIONS(data_list, i + 3) # Option C
                ws[f"F{ROW}"] = EXTRA_OPTIONS(data_list, i + 4) # Option D

                # Increment the row counter by 1
                ROW += 1

            else:
                print(f"Invalid data format at index {i}")
        # Save the workbook as an Excel file
        wb.save(_XLSX_FILENAME)
        CHECK_EMPTY_CELL(_XLSX_FILENAME, ROW)

        print('\x1b[6;30;42m' + 'Success!' + '\x1b[0m')
        # return ROW
    except ValueError as e:
        print("An error occurred:", str(e))
