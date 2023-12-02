import openpyxl
import sys

def CHECK_EMPTY_CELL(excel_filename, number_question):
    wb = openpyxl.load_workbook(excel_filename)
    count = 0  # Initialize a counter to keep track of the number of printed cells

    for rowNumber in range(3, number_question):
        if wb.active.cell(row=rowNumber, column=7).value is None:
            print(f"Miss answer: {rowNumber - 2} Question: {wb.active.cell(row=rowNumber, column=1).value}")
            count += 1  # Increment the counter
            if count >= 10:
                break  # Exit the loop after printing 10 cells

        elif wb.active.cell(row=rowNumber, column=1).value is None:
            print(f"Miss question: {rowNumber - 2}")
            count += 1  # Increment the counter
            if count >= 10:
                break  # Exit the loop after printing 10 cells