"""
This function writes data to an excel file.

The function imports the following modules:

* `openpyxl`: This module provides a Python interface to Excel files.
* `module.extractor`: This module contains functions to extract data from a docx file.

The `write_to_excel()` function has the following parameters:

* `data_list`: A list of data to be written to the excel file.
* `excel_filename`: The path to the excel file.

The `write_to_excel()` function uses the following steps:

1. Opens an excel workbook.
2. Creates a worksheet object.
3. Initializes a row counter.
4. Iterates over the data list with a step of 5.
5. Checks if the data list has at least 4 elements from index i.
6. If the data list has at least 4 elements from index i, writes the question and its options to the worksheet cells.
7. Increments the row counter by 1.
8. Saves the workbook as an Excel file with the given filename.
"""

from module.extractor import extra_questions
from module.extractor import check_format_op
from module.extractor import extra_options
from openpyxl import load_workbook

def write_to_excel(data_list, excel_filename):
    """
    Writes data to an excel file.

    Args:
        data_list: A list of data to be written to the excel file.
        excel_filename: The path to the excel file.

    Returns:
        The row number of the last row written to the excel file.
    """
    
    # Opens an excel workbook.
    wb = load_workbook(excel_filename)
    # Opens an excel workbook.
    ws = wb.active

    # Initialize a row counter
    row = 3

    # Loop over the data list with a step of 5
    for i in range(0, len(data_list), 5):
        # Check if the data_list has at least 4 elements from index i
        if len(data_list) >= i + 4:
            # Write the question and its options to the worksheet cells
            ws[f"A{row}"] = extra_questions(data_list, i) # The question
            ws[f"B{row}"] = "Multiple Choice"
            ws[f"C{row}"] = extra_options(data_list, i + 1)# Option A
            ws[f"D{row}"] = extra_options(data_list, i + 2) # Option B
            ws[f"E{row}"] = extra_options(data_list, i + 3) # Option C
            ws[f"F{row}"] = extra_options(data_list, i + 4) # Option D

            # Increment the row counter by 1
            row += 1

        else:
            pass
            # messagebox.showerror(title='ERROR', message=str(f"Invalid data format at index {i}"))

    check_format_op()
    # Save the workbook as an Excel file with the given filename
    wb.save(excel_filename)
    return row
